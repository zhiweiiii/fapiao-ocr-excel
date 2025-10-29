import os
import json
from pathlib import Path
from typing import List

from openpyxl import load_workbook
import logging
import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import importlib.util
MAIN_PATH = Path(__file__).resolve().parent.parent / 'main.py'
spec = importlib.util.spec_from_file_location("main", MAIN_PATH)
main_mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(main_mod)
create_invoices_with_pandas = getattr(main_mod, 'create_invoices_with_pandas')

DATA_DIR = Path(__file__).resolve().parent.parent / 'data'
OUTPUT_DIR = Path(__file__).resolve().parent.parent / 'output'

REQUIRED_MAIN_FIELDS = [
    '发票类型', '发票号码', '开票日期',
    '购买方名称', '购买方统一社会信用代码/纳税人识别号',
    '销售方名称', '销售方统一社会信用代码/纳税人识别号',
    '合计金额', '合计税额', '价税合计（大写）', '价税合计（小写）',
]

REQUIRED_ITEM_FIELDS = [
    '项目名称', '规格型号', '单位', '数量', '单价', '金额', '税率/征收率', '税额'
]

def find_test_pairs() -> List[Path]:
    """查找 data 目录下同名的 pdf/json 对。"""
    pairs = []
    for json_path in DATA_DIR.glob('*.json'):
        pdf_path = DATA_DIR / (json_path.stem + '.pdf')
        if pdf_path.exists():
            pairs.append(json_path)
    return pairs


def load_json(path: Path) -> dict:
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)


def assert_sheet_headers(ws, required_headers: List[str]):
    headers = [cell.value for cell in ws[1]]
    for h in required_headers:
        assert h in headers, f'缺少必需表头: {h}'


def assert_main_row(ws, data: dict):
    headers = [cell.value for cell in ws[1]]
    row = [cell.value for cell in ws[2]]
    header_index = {h: i for i, h in enumerate(headers)}

    # 核对关键字段的值
    checks = {
        '发票类型': data.get('invoice_type'),
        '发票号码': data.get('invoice_number'),
        '开票日期': data.get('invoice_date'),
        '购买方名称': data.get('buyer_name'),
        '购买方统一社会信用代码/纳税人识别号': data.get('buyer_tax_id'),
        '销售方名称': data.get('seller_name'),
        '销售方统一社会信用代码/纳税人识别号': data.get('seller_tax_id'),
        '合计金额': data.get('total_amount'),
        '合计税额': data.get('total_tax'),
        '价税合计（大写）': data.get('total_with_tax_cn'),
        '价税合计（小写）': data.get('total_with_tax_num'),
    }
    for header, expected in checks.items():
        assert header in header_index, f'主表缺少字段: {header}'
        idx = header_index[header]
        actual = row[idx]
        assert str(actual) == str(expected), f'{header} 值不匹配: 期望 {expected}, 实际 {actual}'


def assert_detail_rows(ws, items: List[dict]):
    headers = [cell.value for cell in ws[1]]
    assert_sheet_headers(ws, REQUIRED_ITEM_FIELDS)
    # 数据行数量 = items 长度
    assert ws.max_row - 1 == len(items), f'明细行数不匹配: 期望 {len(items)}, 实际 {ws.max_row - 1}'

# 从 main.py 额外获取识别结果解析方法
extract_invoice_info = getattr(main_mod, 'extract_invoice_info')

# 加载 OCR 管理器（thread_single.PaddleOCRModelManager）
THREAD_PATH = Path(__file__).resolve().parent.parent / 'thread_single.py'
spec_thr = importlib.util.spec_from_file_location("thread_single", THREAD_PATH)
thread_mod = importlib.util.module_from_spec(spec_thr)
spec_thr.loader.exec_module(thread_mod)
PaddleOCRModelManager = getattr(thread_mod, 'PaddleOCRModelManager')

class _DummyApp:
    def __init__(self):
        self.logger = logging.getLogger("test")
        if not self.logger.handlers:
            h = logging.StreamHandler()
            fmt = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            h.setFormatter(fmt)
            self.logger.addHandler(h)
        self.logger.setLevel(logging.INFO)

# 计算主表字段准确率
MAIN_FIELD_KEYS = [
    'invoice_type', 'invoice_number', 'invoice_date',
    'buyer_name', 'buyer_tax_id', 'seller_name', 'seller_tax_id',
    'total_amount', 'total_tax', 'total_with_tax_cn', 'total_with_tax_num'
]

def compute_main_accuracy(expected: dict, ocr: dict):
    total = len(MAIN_FIELD_KEYS)
    match = 0
    mismatches = []
    for k in MAIN_FIELD_KEYS:
        ev = str(expected.get(k, '')).strip()
        ov = str(ocr.get(k, '')).strip()
        if ev == ov:
            match += 1
        else:
            mismatches.append((k, ev, ov))
    acc = match / total if total else 0.0
    return acc, mismatches

# 计算明细表字段准确率（按索引对齐）
ITEM_FIELD_KEYS = ['product_name', 'specification', 'unit', 'quantity', 'unit_price', 'amount', 'tax_rate', 'tax_amount']

def compute_items_accuracy(expected_items: list, ocr_items: list):
    n = min(len(expected_items), len(ocr_items))
    total = n * len(ITEM_FIELD_KEYS) if n else 0
    match = 0
    mismatches = []
    for i in range(n):
        eitem = expected_items[i]
        oitem = ocr_items[i]
        for k in ITEM_FIELD_KEYS:
            ev = str(eitem.get(k, '')).strip()
            ov = str(oitem.get(k, '')).strip()
            if ev == ov:
                match += 1
            else:
                mismatches.append((i, k, ev, ov))
    acc = match / total if total else 0.0
    return acc, mismatches

def run_case(json_path: Path):
    print(f'运行用例: {json_path.stem}')
    data = load_json(json_path)
    assert 'items' in data and isinstance(data['items'], list), 'JSON中缺少 items 列表'

    # 调用 OCR 识别 pdf
    pdf_path = DATA_DIR / (json_path.stem + '.pdf')
    assert pdf_path.exists(), f'未找到对应的PDF: {pdf_path}'

    ocr_manager = PaddleOCRModelManager(_DummyApp())
    _, result_all = ocr_manager.submit_ocr(input=str(pdf_path))
    ocr_list = extract_invoice_info(result_all)
    assert ocr_list, 'OCR未解析出发票信息'
    ocr_info = ocr_list[0]

    # 计算准确率（主表 + 明细）
    main_acc, main_mismatches = compute_main_accuracy(data, ocr_info)
    items_acc, items_mismatches = compute_items_accuracy(data.get('items', []), ocr_info.get('items', []))
    # 综合准确率按字段总数加权
    main_total = len(MAIN_FIELD_KEYS)
    items_total = min(len(data.get('items', [])), len(ocr_info.get('items', []))) * len(ITEM_FIELD_KEYS)
    total_fields = main_total + items_total
    total_match = int(main_acc * main_total) + int(items_acc * (items_total if items_total else 0))
    overall_acc = (total_match / total_fields) if total_fields else 0.0

    print(f'主表准确率: {main_acc*100:.2f}%  不匹配: {len(main_mismatches)}')
    for k, ev, ov in main_mismatches:
        print(f'  字段[{k}] 期望={ev} 识别={ov}')
    print(f'明细准确率: {items_acc*100:.2f}%  不匹配: {len(items_mismatches)}')
    for i, k, ev, ov in items_mismatches[:10]:  # 仅打印前10项避免日志过长
        print(f'  第{i+1}行字段[{k}] 期望={ev} 识别={ov}')
    print(f'用例[{json_path.stem}] 综合准确率: {overall_acc*100:.2f}%')

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_xlsx = OUTPUT_DIR / f'test_{json_path.stem}.xlsx'

    # 生成Excel
    created_path = create_invoices_with_pandas([data], output_path=str(out_xlsx))
    assert Path(created_path).exists(), f'Excel未生成: {created_path}'

    # 校验Excel内容
    wb = load_workbook(created_path)
    assert '发票主表' in wb.sheetnames, '缺少工作表: 发票主表'
    assert '发票明细' in wb.sheetnames, '缺少工作表: 发票明细'

    ws_main = wb['发票主表']
    ws_detail = wb['发票明细']
    assert_sheet_headers(ws_main, REQUIRED_MAIN_FIELDS)
    assert_main_row(ws_main, data)
    assert_detail_rows(ws_detail, data['items'])
    print(f'用例通过: {json_path.stem}')


if __name__ == '__main__':
    pairs = find_test_pairs()
    assert pairs, f'在 {DATA_DIR} 未找到任何 pdf/json 同名对（至少应存在一个 .json 与同名 .pdf）'
    accs = []
    for json_path in pairs:
        accs.append(run_case(json_path))
    avg_acc = sum(accs)/len(accs) if accs else 0.0
    print(f'全部用例平均准确率: {avg_acc*100:.2f}%')
    print('全部用例完成')